from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import load_workbook


class DatasetError(ValueError):
    """Raised when an uploaded dataset cannot be safely parsed."""


@dataclass(frozen=True)
class DatasetLimits:
    max_files: int = 10
    max_file_size_bytes: int = 10 * 1024 * 1024
    max_model_rows: int = 5000
    max_sample_rows: int = 8


class DatasetService:
    """Parse uploaded datasets into one provider-neutral modeling contract.

    The service owns file-format concerns only. It does not infer joins,
    optimization variables, objectives, or constraints; those remain the
    modeling engine's responsibility.
    """

    SUPPORTED_FORMATS = {".csv", ".txt", ".xlsx"}

    def __init__(self, limits: DatasetLimits | None = None) -> None:
        self.limits = limits or DatasetLimits()

    def summarize_files(self, files: Sequence[tuple[str, bytes]]) -> Dict[str, Any]:
        if not files:
            raise DatasetError("At least one dataset is required.")
        if len(files) > self.limits.max_files:
            raise DatasetError(f"A maximum of {self.limits.max_files} datasets can be uploaded.")

        sources: List[Dict[str, Any]] = []
        total_size = 0
        for filename, content in files:
            source = self.summarize_file(content, filename)
            sources.extend(source if isinstance(source, list) else [source])
            total_size += len(content)

        # Keep a flat legacy shape for the single-source path. The LLM and
        # deterministic tabular completion can therefore retain their existing
        # behavior without making multi-source parsing part of the domain layer.
        result: Dict[str, Any] = {
            "source_count": len(sources),
            "sources": sources,
            "multi_source": len(sources) > 1,
            "total_size_bytes": total_size,
        }
        if len(sources) == 1:
            result.update({
                key: sources[0].get(key)
                for key in ("filename", "format", "row_count", "column_count", "columns", "sample_rows", "rows", "rows_truncated")
            })
        return result

    def summarize_file(self, content: bytes, filename: str) -> Dict[str, Any] | List[Dict[str, Any]]:
        safe_name = Path(filename or "dataset").name
        suffix = Path(safe_name).suffix.lower()
        if suffix not in self.SUPPORTED_FORMATS:
            supported = ", ".join(sorted(self.SUPPORTED_FORMATS))
            raise DatasetError(f"Unsupported dataset format '{suffix or 'unknown'}'. Supported formats: {supported}.")
        if len(content) > self.limits.max_file_size_bytes:
            raise DatasetError(f"'{safe_name}' is too large. Maximum size is 10 MB per file.")

        if suffix == ".xlsx":
            return self._summarize_xlsx(content, safe_name)
        if suffix == ".txt":
            return self._summarize_delimited_or_text(content, safe_name, suffix)
        return self._summarize_delimited_or_text(content, safe_name, suffix)

    def _decode_text(self, content: bytes, filename: str) -> str:
        try:
            return content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise DatasetError(f"'{filename}' must be encoded as UTF-8.") from exc

    def _summarize_delimited_or_text(self, content: bytes, filename: str, suffix: str) -> Dict[str, Any]:
        text = self._decode_text(content, filename)
        if not text.strip():
            raise DatasetError(f"'{filename}' is empty.")

        # CSV is strict. TXT is intentionally flexible: a delimited text file
        # becomes a table, while an ordinary text file becomes one text column.
        if suffix == ".txt":
            try:
                sample = "\n".join(text.splitlines()[:20])
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                has_header = csv.Sniffer().has_header(sample)
            except csv.Error:
                return self._summarize_plain_text(text, filename)
            if not has_header:
                return self._summarize_plain_text(text, filename)
            delimiter = dialect.delimiter
        else:
            delimiter = None
            try:
                sample = "\n".join(text.splitlines()[:20])
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ","

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        if not reader.fieldnames:
            raise DatasetError(f"'{filename}' must contain a header row.")
        return self._summarize_reader(reader, filename, suffix.lstrip("."))

    def _summarize_plain_text(self, text: str, filename: str) -> Dict[str, Any]:
        lines = text.splitlines()
        rows = [{"line_number": index, "text": line} for index, line in enumerate(lines, start=1)]
        return {
            "filename": filename,
            "format": "txt",
            "row_count": len(rows),
            "column_count": 2,
            "columns": ["line_number", "text"],
            "sample_rows": rows[: self.limits.max_sample_rows],
            "rows": rows[: self.limits.max_model_rows],
            "rows_truncated": len(rows) > self.limits.max_model_rows,
            "source_kind": "text",
        }

    def _summarize_reader(self, reader: Any, filename: str, fmt: str) -> Dict[str, Any]:
        columns = [str(name).strip() for name in reader.fieldnames or [] if name and str(name).strip()]
        if not columns:
            raise DatasetError(f"'{filename}' must contain at least one named column.")

        rows: List[Dict[str, Any]] = []
        sample_rows: List[Dict[str, Any]] = []
        row_count = 0
        for row in reader:
            row_count += 1
            normalized = {column: row.get(column) for column in columns}
            if len(sample_rows) < self.limits.max_sample_rows:
                sample_rows.append(normalized)
            if len(rows) < self.limits.max_model_rows:
                rows.append(normalized)

        return {
            "filename": filename,
            "format": fmt,
            "row_count": row_count,
            "column_count": len(columns),
            "columns": columns,
            "sample_rows": sample_rows,
            "rows": rows,
            "rows_truncated": row_count > self.limits.max_model_rows,
            "source_kind": "tabular",
        }

    def _summarize_xlsx(self, content: bytes, filename: str) -> List[Dict[str, Any]]:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise DatasetError(f"Could not read '{filename}' as XLSX: {exc}") from exc

        sources: List[Dict[str, Any]] = []
        try:
            for sheet in workbook.worksheets:
                values = sheet.iter_rows(values_only=True)
                try:
                    header = next(values)
                except StopIteration:
                    continue
                columns = [str(value).strip() if value is not None else "" for value in header]
                if not any(columns):
                    continue
                columns = [value or f"column_{index + 1}" for index, value in enumerate(columns)]

                rows: List[Dict[str, Any]] = []
                sample_rows: List[Dict[str, Any]] = []
                row_count = 0
                for raw_row in values:
                    normalized = {
                        column: raw_row[index] if index < len(raw_row) else None
                        for index, column in enumerate(columns)
                    }
                    # Ignore completely empty spreadsheet rows.
                    if all(value is None or str(value).strip() == "" for value in normalized.values()):
                        continue
                    row_count += 1
                    if len(sample_rows) < self.limits.max_sample_rows:
                        sample_rows.append(normalized)
                    if len(rows) < self.limits.max_model_rows:
                        rows.append(normalized)

                sources.append({
                    "filename": filename,
                    "sheet": sheet.title,
                    "source_name": f"{filename}::{sheet.title}",
                    "format": "xlsx",
                    "row_count": row_count,
                    "column_count": len(columns),
                    "columns": columns,
                    "sample_rows": sample_rows,
                    "rows": rows,
                    "rows_truncated": row_count > self.limits.max_model_rows,
                    "source_kind": "tabular",
                })
        finally:
            workbook.close()

        if not sources:
            raise DatasetError(f"'{filename}' does not contain any non-empty worksheet with a header row.")
        return sources
